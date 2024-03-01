# import requests
# from bs4 import BeautifulSoup as BS
# r = requests.get("https://stopgame.ru/news/all/p1")
# html = BS(r.content, 'html.parser')

# for el in html.select(".list-view > ._card_1tbpr_1"):
#     title = el.select('.caption > a')
#     print(title[0].text)

# from openpyxl import Workbook
# work_book = Workbook()
# work_sheet = work_book.active

# # Коллекция для генерации таблицы
# numbers = [1, 2, 3, 4, 5]

# for elem in numbers:
#     row = [elem, elem * 2, elem ** 2]
#     print(row)
#     work_sheet.append(row)

# work_book.save('numbers.xlsx')



# from openpyxl import Workbook
# work_book = Workbook()
# work_sheet = work_book.active

# import requests
# web_page = requests.get('https://live.skillbox.ru/playlists/code/python/')
# web_page.text

# from bs4 import BeautifulSoup
# soup = BeautifulSoup(web_page.text, 'html.parser')
# items = soup.find_all(class_='playlist-inner__item')

# for elem in items:
#     title = elem.find(class_='playlist-inner-card__link-text').text
#     relative_url = elem.find(class_='playlist-inner-card__link').attrs['href']
#     url = 'https://live.skillbox.ru' + relative_url
#     row = [title, url]
#     print(row)
#     work_sheet.append(row)

# work_book.save('Вебинары про Python от Skillbox.xlsx')
print('Миша')
from openpyxl import Workbook
work_book = Workbook()
work_sheet = work_book.active


import requests
#web_page = requests.get('https://www.ranepa.ru/press-sluzhba/news/')
# web_page.text
from bs4 import BeautifulSoup as bs
r = requests.get('https://мвд.рф/?%EF=')
#print(r)
soup = bs(r.content, 'html.parser')
#print(soup)
listof = soup.find_all(class_="news-popup e-popup")
for s in listof:
    print(s.attrs['href'].attrs['rel'])
    #print(s.attrs['rel'])
print(len(listof))
exit()
#soup = BeautifulSoup(web_page.content, 'html.parser')
# r = requests.get("https://www.ranepa.ru/press-sluzhba/news/")
r = requests.get('https://www.ranepa.ru/')
print(r)
soup = bs(r.content, 'html.parser')
print(soup)
# print(soup.find(class_='news-list-head').text)
# exit()
# a = soup.find_all(class_='standartnews__title')
# print(len(a))
# for i in soup.find_all(class_='access'):
#    print(i)
print(soup.find(class_="col-xs-12").text)
listof = soup.find(class_='container branches').find_all(class_='row')
for headline in listof:
    row1 = headline.find(class_='branch__title').attrs['itemprop'].text
    row2 = headline.find(class_='branch__text branch__text--adress').attrs['itemprop'].text
    print(row1, row2)
    work_sheet.append((row1, row2))
work_book.save('Hi Sanya')